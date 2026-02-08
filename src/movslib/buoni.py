from datetime import UTC
from datetime import datetime
from decimal import Decimal
from operator import attrgetter
from typing import TYPE_CHECKING
from typing import Final
from typing import overload

from openpyxl import load_workbook

from movslib.model import KV
from movslib.model import Row
from movslib.model import Rows

if TYPE_CHECKING:
    from collections.abc import Iterable

    from openpyxl.worksheet.worksheet import Worksheet


SHEET_NAME: Final = 'RPOL_PatrimonioBuoni'
MIN_ROW: Final = 2
MAX_ROW: Final = 999


def _load_sheet(fn: str) -> 'Worksheet':
    wb = load_workbook(fn, read_only=True, data_only=True, keep_links=False)
    return wb[SHEET_NAME]


def _read_csv(sheet: 'Worksheet') -> 'Iterable[Row]':
    for (
        _,
        tipologia,
        _,
        _,
        _,
        valore_rimborso_netto_raw,
        _,
        _,
        _,
        data_sottoscrizione_raw,
        valore_nominale_raw,
        scadenza_raw,
        serie,
        regolato_su,
        *_,
    ) in sheet.iter_rows(
        MIN_ROW, MAX_ROW, min_col=1, max_col=19, values_only=True
    ):
        if not isinstance(tipologia, str):
            raise TypeError(type(tipologia))
        if not isinstance(valore_rimborso_netto_raw, str):
            raise TypeError(type(valore_rimborso_netto_raw))
        if not isinstance(data_sottoscrizione_raw, str):
            raise TypeError(type(data_sottoscrizione_raw))
        if not isinstance(valore_nominale_raw, str):
            raise TypeError(type(valore_nominale_raw))
        if not isinstance(scadenza_raw, str):
            raise TypeError(type(scadenza_raw))
        if not isinstance(serie, str):
            raise TypeError(type(serie))
        if not isinstance(regolato_su, str):
            raise TypeError(type(regolato_su))

        valore_rimborso_netto = Decimal(
            valore_rimborso_netto_raw[1:].replace('.', '').replace(',', '.')
        )
        data_sottoscrizione = (
            datetime.strptime(data_sottoscrizione_raw[:10], '%d/%m/%Y')
            .replace(tzinfo=UTC)
            .date()
        )
        valore_nominale = Decimal(
            valore_nominale_raw[1:].replace('.', '').replace(',', '.')
        )
        scadenza = (
            datetime.strptime(scadenza_raw, '%d.%m.%Y')
            .replace(tzinfo=UTC)
            .date()
        )

        yield Row(
            data_contabile=data_sottoscrizione,
            data_valuta=data_sottoscrizione,
            addebiti=valore_nominale,
            accrediti=None,
            descrizione_operazioni=f'sottoscrizione {tipologia},{serie},{regolato_su}',
        )
        yield Row(
            data_contabile=scadenza,
            data_valuta=scadenza,
            addebiti=None,
            accrediti=valore_rimborso_netto,
            descrizione_operazioni=f'rimborso {tipologia},{serie},{regolato_su}',
        )


def read_csv(fn_sheet: 'str | Worksheet') -> 'Iterable[Row]':
    sheet = _load_sheet(fn_sheet) if isinstance(fn_sheet, str) else fn_sheet
    return sorted(_read_csv(sheet), key=attrgetter('date'), reverse=True)


@overload
def read_buoni(fn: str) -> 'tuple[KV, list[Row]]': ...


@overload
def read_buoni(fn: str, name: str) -> 'tuple[KV, Rows]': ...


def read_buoni(
    fn: str, name: str | None = None
) -> 'tuple[KV, list[Row] | Rows]':
    sheet = _load_sheet(fn)
    csv = read_csv(sheet)
    kv = KV(
        da=None,
        a=None,
        tipo='buoni postali',
        conto_bancoposta='',
        intestato_a='',
        saldo_al=None,
        saldo_contabile=sum(row.money for row in csv),
        saldo_disponibile=sum(row.money for row in csv),
    )

    return kv, (list(csv) if name is None else Rows(name, csv))
